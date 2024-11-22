import os
from datetime import datetime
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import PatternFill


def export_data_to_excel(data_list, purposes):
    """
    Export multiple experiment data objects to an Excel file, each saved to a separate sheet.

    Parameters:
        data_list (list): A list of dictionaries containing the experiment data.
                          Each dictionary should have keys: 'creator', 'experimentName', 'matrixSize', etc.
        purposes (str): The purpose of the export (e.g., 'feedback').

    Returns:
        tuple: The filename and absolute path of the generated Excel file.
    """
    # Create folder for saving files
    folder_name = "Feedback Data" if purposes == 'feedback' else "Experiment Data"

    if not os.path.exists(folder_name):
        os.makedirs(folder_name)

    # Generate the filename with timestamp
    filename = os.path.join(
        folder_name, datetime.now().strftime('%Y-%m-%d %H' + "h" + '%M' + 'p') + '.xlsx'
    )

    # Create a workbook
    wb = Workbook()
    wb.remove(wb.active)  # Remove the default sheet created by openpyxl

    # Iterate over each data object in the list
    for data in data_list:
        # Extract details from the input data
        creator = data.get('creator', 'Creator')
        experiment_name = data.get('experimentName', 'Experiment')
        stageNumber = data.get('stageNumber', 'Stage Number')
        sheet_name = f"Stage {stageNumber}" if stageNumber else "Sheet"

        # Prepare the data based on purposes
        if purposes == 'feedback':
            lor = data.get('levelOfRecognition', 'Level of Recognition')
            config = data.get('config', [])
            matrixGrid = data.get('matrixGrid', [])

            # Convert config to a DataFrame
            df = pd.DataFrame(config, columns=[
                'Index', 'Node_number', 'Intensity', 'Duration', 'Order'])

            # Convert matrixGrid to a DataFrame
            df2 = pd.DataFrame(matrixGrid, columns=[
                'index', 'nodeNumber', 'pressedCount'])

            # Combine columns
            df3 = pd.concat([df, df2], axis=1)

            # Add additional columns with values copied if 'pressedCount' is not empty
            df3['Level of Recognition'] = df3['pressedCount'].apply(lambda x: lor if pd.notnull(x) else '')
            df3['Stage Number'] = df3['pressedCount'].apply(lambda x: stageNumber if pd.notnull(x) else '')

        else:
            records = data.get('records', [])
            df3 = pd.DataFrame(records, columns=[
                'Index', 'Node_number', 'Intensity', 'Duration', 'Order'])

        # Add the DataFrame to a new sheet in the workbook
        ws = wb.create_sheet(title=sheet_name)

        # Add headers to the sheet
        for col_idx, header in enumerate(df3.columns, start=1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            # Apply yellow color for headers
            if col_idx <= 5:  # Yellow for the first 5 columns
                cell.fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
            else:  # Blue for the rest
                cell.fill = PatternFill(start_color="87CEEB", end_color="87CEEB", fill_type="solid")

        # Add data rows to the sheet
        for r_idx, row in enumerate(df3.itertuples(index=False), start=2):
            for c_idx, value in enumerate(row, start=1):
                cell = ws.cell(row=r_idx, column=c_idx, value=value)
                # Apply yellow color for data
                if c_idx <= 5:  # Yellow for the first 5 columns
                    cell.fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
                else:  # Blue for the rest
                    cell.fill = PatternFill(start_color="87CEEB", end_color="87CEEB", fill_type="solid")

    # Save the workbook
    wb.save(filename)

    # Return the filename and absolute path
    return filename, os.path.abspath(filename)
